import csv
import re
import os
import time
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


BANK_STATEMENTS_FOLDER_PATH = "CSV_bank_statements"

MONTHS = {
    1: 'January',
    2: 'February',
    3: 'March',
    4: 'April',
    5: 'May',
    6: 'June',
    7: 'July',
    8: 'August',
    9: 'September',
    10: 'October',
    11: 'November',
    12: 'December'}

HIGHLIGHT_YELLOW = PatternFill(start_color="FFFF00",
                               end_color="FFFF00", fill_type="solid")

HIGHLIGHT_ORANGE = PatternFill(start_color="FFA500",
                               end_color="FFA500", fill_type="solid")

HIGHLIGHT_RED = PatternFill(start_color="FF0000",
                            end_color="FF0000", fill_type="solid")

wb_category = xl.load_workbook("Categories.xlsx")
ws_category = wb_category["Sheet1"]


def bank_statements_folder(FOLDER_PATH):
    for filename in os.listdir(FOLDER_PATH):
        file_path = os.path.join(FOLDER_PATH, filename)
        if os.path.isfile(file_path):
            yield (file_path, filename)


def categorize(line):
    col = 65

    while ws_category[chr(col)+"1"].value != None:
        char = chr(col)
        row = 2

        while ws_category[char + str(row)].value != None:
            cell_val = ws_category[char + str(row)].value
            pattern = re.compile(f"{cell_val.lower().strip()}")
            match = pattern.search(line.lower())

            if match:
                return ws_category[char + "1"].value

            row += 1
        col += 1

    else:
        return "ASK"


def highlight(ws):
    row = 2

    while ws["D"+str(row)].value != None:
        cell = ws.cell(row, 4)

        if cell.value == "ASK":
            cell.fill = HIGHLIGHT_RED

        if cell.value == "Car Expenses":
            if ws["B"+str(row)].value < 10:
                cell.value = "Meals/Entertainment"

        elif cell.value == "Check":
            if ws["B"+str(row)].value > 2500:
                cell.value = "Material"
                cell.fill = HIGHLIGHT_ORANGE

            else:
                cell.value = "Labor"
                cell.fill = HIGHLIGHT_YELLOW

        row += 1


def main():

    for csv_file, name in bank_statements_folder(BANK_STATEMENTS_FOLDER_PATH):
        wb = xl.Workbook()

        with open(csv_file, newline="") as current_file:
            fieldnames = ["Date", "Amount", "Star", "Check#", "Description"]

            csv_reader = csv.DictReader(current_file, fieldnames=fieldnames)

            for i in range(1, 13):
                ws = wb.create_sheet(title=MONTHS[i])
                ws.append(["Date", "Amount", "Description", "Category"])

                for line in csv_reader:
                    date = line["Date"]
                    amount = float(line["Amount"])

                    if int(date.split("/")[0]) == i and amount < 0:
                        description = line["Description"]
                        category = categorize(description)
                        ws.append([date, abs(amount), description, category])

                highlight(ws)
                current_file.seek(0)

        wb.save(f"excel_files/{name}.xlsx")


if __name__ == "__main__":
    start = time.time()
    main()
    end = time.time()
    print(f"{end - start:.3f} seconds")
