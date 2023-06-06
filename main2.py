import pdfplumber
import re
import os
import time
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


BANK_STATEMENTS_FOLDER_PATH = "bank_statements"
MONTHS = {
    1: 'January',
    2: 'February',
    3: 'March',
    4: 'April',
    5: 'May',
    6: 'June',
    7: 'July',
    8: 'August',
    9: 'September',
    10: 'October',
    11: 'November',
    12: 'December'}

HIGHLIGHT_YELLOW = PatternFill(start_color="FFFF00",
                               end_color="FFFF00", fill_type="solid")

HIGHLIGHT_ORANGE = PatternFill(start_color="FFA500",
                               end_color="FFA500", fill_type="solid")

HIGHLIGHT_RED = PatternFill(start_color="FF0000",
                            end_color="FF0000", fill_type="solid")


def statements_finished(line):
    pattern = re.compile(r"^Ending")
    match = pattern.search(line)
    return match


# yields files in folder
def bank_statements_folder(FOLDER_PATH):
    for filename in os.listdir(FOLDER_PATH):
        file_path = os.path.join(FOLDER_PATH, filename)
        if os.path.isfile(file_path):
            yield file_path


# searches for dates at the beggining of each line
def find_date(line):
    pattern = re.compile(r"^\d+/\d+")
    match = pattern.search(line)
    return match


# gets amount of money involved in transaction
def transaction_amount(line):
    pattern = re.compile(r" [\d,?]+[.]\d\d")
    match = pattern.search(line)
    return match


def deposit(line):
    pattern = re.compile(
        r"Deposit |Square Inc|Online Transfer From|Card Provisional Credit|Edeposit|RTP From Square")
    return pattern.search(line)


def bank_statement_file(file):
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages[1:]:
            for line in page.extract_text().split("\n"):
                yield line


def set_month(line, ws):
    month = int(line.split("/")[0])
    ws.title = MONTHS[month]


def appends_row_income(line, description, income_count, ws):
    row = []
    row.append(find_date(line).group(0))  # date
    row.append(transaction_amount(line).group(0))
    row.append(description)

    for col in range(3):  # loops from column A-C
        char = get_column_letter(col+1)
        ws[char + str(income_count+3)] = row[col]


def appends_row_expense(line, description, category, expenses_count, ws):
    row = []
    row.append(find_date(line).group(0))  # date
    row.append(transaction_amount(line).group(0))
    row.append(description)
    row.append(category)

    for col in range(4):  # loops from column G-I
        char = get_column_letter(col+7)
        cell = ws[char + str(expenses_count+3)]
        amount = float(transaction_amount(line).group(0).replace(",", ""))

        if col == 3:
            if row[col] == "ASK":
                cell.fill = HIGHLIGHT_RED

            if row[col] == "Car Expenses":
                if amount < 10:
                    row[3] = "Meals/Entertainment"

            elif row[col] == "Check":
                if float(transaction_amount(line).group(0).replace(",", "")) > 2500:
                    row[3] = "Material"
                    cell.fill = HIGHLIGHT_ORANGE

                else:
                    row[3] = "Labor"
                    cell.fill = HIGHLIGHT_YELLOW

        ws[char + str(expenses_count+3)] = row[col]


def clean_description(line):
    strings = line.split()
    pattern = re.compile(r"[\d,?]+[.]\d\d")
    description = []

    if strings[1] == "Purchase":
        for item in strings[5:]:
            if not pattern.search(item):
                description.append(item)
        return " ".join(description)

    elif strings[1] == "Recurring":
        for item in strings[6:]:
            if not pattern.search(item):
                description.append(item)
        return " ".join(description)

    elif strings[1] == "ATM":
        return "ATM Withdrawal"

    elif strings[2] == "Check" or strings[2] == "Cashed" or strings[2] == "Deposited":
        return "Cashed Check"

    else:
        for item in strings[1:]:
            if not pattern.search(item):
                description.append(item)
        return " ".join(description)


def clean_description_income(line):
    strings = line.split()
    pattern = re.compile(r"[\d,?]+[.]\d\d")
    description = []

    if strings[1] == "Mobile":
        return "Mobile Deposit"

    elif strings[1] == "ATM":
        return "ATM Deposit"

    elif strings[1] == "Square":
        for item in strings[6:]:
            if not pattern.search(item):
                description.append(item)
        return " ".join(description)

    elif strings[1] == "Online":
        for item in strings[4:]:
            if not pattern.search(item):
                description.append(item)
        return " ".join(description)

    else:
        for item in strings[1:]:
            if not pattern.search(item):
                description.append(item)
        return " ".join(description)


def categorize(line):
    wb = xl.load_workbook("Categories.xlsx")
    ws = wb["Sheet1"]
    col = 65

    while ws[chr(col)+"1"].value != None:
        char = chr(col)
        row = 2

        while ws[char + str(row)].value != None:
            cell_val = ws[char + str(row)].value
            pattern = re.compile(f"{cell_val.lower().strip()}")
            match = pattern.search(line.lower())

            if match:
                return ws[char + "1"].value

            row += 1
        col += 1

    else:
        return "ASK"


def main():
    wb = xl.Workbook()

    for file in bank_statements_folder(BANK_STATEMENTS_FOLDER_PATH):
        ws = wb.create_sheet()
        ws.append(["Income"])
        ws.merge_cells("A1:D1")
        ws["G1"] = "Expenses"
        ws.merge_cells("G1:J1")
        ws.append(["Date", "Amount", "Description", "Category",
                  "", "", "Date", "Amount", "Description", "Category"])
        expenses_count = 0
        income_count = 0
        statements_looping = False
        month_not_found = True

        for line in bank_statement_file(file):

            if statements_looping:

                if statements_finished(line):
                    break

            if find_date(line):
                statements_looping = True

                if month_not_found:
                    set_month(line, ws)
                    month_not_found = False

                if deposit(line):
                    description = clean_description_income(line)
                    appends_row_income(line, description, income_count, ws)
                    income_count += 1

                else:
                    description = clean_description(line)
                    category = categorize(line)
                    appends_row_expense(line, description,
                                        category, expenses_count, ws)
                    expenses_count += 1

    wb.save("Monthly Expenses.xlsx")


if __name__ == "__main__":
    start = time.time()
    main()
    end = time.time()
    print(f"{end - start:.3f} seconds")
